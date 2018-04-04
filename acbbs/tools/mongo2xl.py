#!/usr/bin/python2.7
# coding=UTF-8

from pymongo import MongoClient
from openpyxl import Workbook, load_workbook
from os import path

class xlm(object):
    def __init__(self, filename):
        self.filename = filename
        if not path.isfile(self.filename):
            self.book = Workbook()
        else:
            self.book = load_workbook(self.filename)
        self.sheet = self.book.active

    def save(self):
        self.book.save(self.filename)

    def write_row(self, row):
        self.sheet.append(row)

    def create_tab(self, tab):
        self.sheet = self.book.create_sheet(tab)


class mongo(object):
    def __init__(self):
        self.client = MongoClient("127.0.0.1", 27017, maxPoolSize=50)
        self.db = self.client["acbbs"]

    def list_collections(self):
        list = []
        for collection in self.db.collection_names():
            if "system.indexes" not in collection:
                list.append(collection)
        return list

    def get_collection(self, collection):
        return self.db[collection].find({})

if __name__ == "__main__":
    pm = mongo()
    xl = xlm(filename="test.xlsx")

    for tc in pm.list_collections():
        print tc
        xl.create_tab(tab=tc)
        title = False

        for document in pm.get_collection(tc):
            row = []
            
            #get title name
            if title is False:
                title = True
                for keys in document["input-parameters"].keys():
                    row.append(keys)
                for keys in document["dut-result"].keys():
                    row.append(keys)
                for keys in document["dut-info"]["measure"].keys():
                    row.append(keys)
                xl.write_row(row)
                row = []

            #get value
            for values in document["input-parameters"].values():
                row.append(values)
            for values in document["dut-result"].values():
                row.append(values)
            for values in document["dut-info"]["measure"].values():
                row.append(values)

            xl.write_row(row)
        xl.save()
